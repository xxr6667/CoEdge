import numpy as np
import pandas as pd
import random
import math
from openpyxl import load_workbook
from scipy.optimize import minimize

### CoEdge: The minimum service delay Min_H when 35-ESP and 55-ESP take on all user requests can be calculated.
### At this time, the scheduling matrix and the service delay of each ESP can also be obtained.

#Take 35-ESP as an example.

# read dataset from task scheduling_35.xlsx
book=load_workbook(r'D:\my work\task scheduling_35.xlsx')

# Distinguish whether ESP requires scheduling or not.
schedule_sheet = book["RS"]
schedule_resource = []
sch_users = []
sch_local = []
surplus_sheet = book["NS"]
surplus_resource = []
sur_users = []

sch_row_num = 2
while sch_row_num <= 13:
    schedule_resource.append(schedule_sheet.cell(row=sch_row_num, column=5).value)
    sch_users.append(schedule_sheet.cell(row=sch_row_num, column=4).value)
    sch_local.append(schedule_sheet.cell(row=sch_row_num, column=2).value)
    sch_row_num = sch_row_num + 1
sur_row_num = 2
while sur_row_num <= 24:
    surplus_resource.append(surplus_sheet.cell(row=sur_row_num, column=5).value)
    sur_users.append(surplus_sheet.cell(row=sur_row_num, column=4).value)
    sur_row_num = sur_row_num + 1
m=len(schedule_resource)
n=len(surplus_resource)
# Link Transmission Rate
R= 0.04
# Computing Latency
sch_computing_delay = [random.uniform(1,30) for _ in range(m)]
sur_computing_delay = [random.uniform(1,30) for _ in range(n)]
# Upload/Download Transmission Latency
sch_updown_delay = [random.uniform(1,30) for _ in range(m)]
sur_updown_delay = [random.uniform(1,30) for _ in range(n)]


# Formulate the objective function and constraints, 
# which include selecting the ESP with the least transmission latency based on the Lagrangian function 
# and labeling the selected ESP as used and update its resources.
def objective_function(x):
   sch_delay=0
   for i in range(m):
        max=x[i*n]*(2*x[i*n]/R+2*sur_updown_delay[0]+sur_computing_delay[0])
        for j in range(n):
            if x[i*n+j]*(2*x[i*n+j]/R+2*sur_updown_delay[j]+sur_computing_delay[j])>max:
                max=x[i*n+j]*(2*x[i*n+j]/R+2*sur_updown_delay[j]+sur_computing_delay[j])
        sch_delay += sch_local[i]*(sch_computing_delay[i]+2*sch_updown_delay[i])+max
   return sch_delay
   
def inequality_constraints(x):
   constraints=[]
   sur_constraints = [surplus_resource[i]-np.sum(x[i::n]) for i in range(n)]#大于等于
   constraints.extend(sur_constraints)
   for i in range(m):
        max=x[i*n]*sur_computing_delay[0]
        for j in range(n):
            if x[i*n+j]*sur_computing_delay[j]>max:
                max=x[i*n+j]*sur_computing_delay[j]
        constraints.append(sch_local[i]*sch_computing_delay[i]-max) 
                           
   return constraints


def equality_constraint_1(x):
   return np.sum(x) - sum(schedule_resource)

def equality_constraint_2(x):
    eq_constraints=[]
    for i in range(m):
       sum=0
       for j in range(n):
           sum+=x[i*n+j]
       eq_constraints.append(sum-schedule_resource[i])
    return eq_constraints

constraints = []
for constraint in inequality_constraints([0]*m*n):
   constraints.append({'type': 'ineq', 'fun': lambda x, c=constraint: c})
constraints.append({'type': 'eq', 'fun': equality_constraint_1})
constraints.append({'type': 'eq', 'fun': equality_constraint_2})
# Define the boundary conditions (non-negative constraints) of the variables.
bounds = [(0, None)] * (m*n)
 
# Number of iterations: 1000
# Convergence tolerance: Convergence is achieved when the relative change is less than 1%.
options = {'disp': True, 'maxiter': 1000, 'ftol': 1e-2}
# Optimization method: Sequential Least Squares Programming (SLSQP)
result = minimize(objective_function, [0]*m*n, method='SLSQP',#'trust-constr',
                 bounds=bounds, constraints=constraints,options=options)

if result.success:
   print(f"minimum value of the objective function: {result.fun}")
   r=np.array(result.x).reshape(m,n)
   df = pd.DataFrame(r).astype('float')
   df.to_excel(r'D:\my work\scheduling matrix_35.xlsx', index=False, engine='openpyxl')
else:
   print("Optimization failed.")
   print("Reasons for failure: ", result.message)


# Calculate the total sevice delay of the RS-ESPs.
r_sch_delay=[]
for i in range(m):
    sch_delay=0
    max=r[i][0]*(2*r[i][0]/R+2*sur_updown_delay[0]+sur_computing_delay[0])
    for j in range(n):
        if r[i][j]*(2*r[i][j]/R+2*sur_updown_delay[j]+sur_computing_delay[j])>max:
            max=r[i][j]*(2*r[i][j]/R+2*sur_updown_delay[j]+sur_computing_delay[j])
    sch_delay += sch_local[i]*(sch_computing_delay[i]+2*sch_updown_delay[i])+max
    r_sch_delay.append(sch_delay)
print(f"total sevice delay of the RS-ESPs: {r_sch_delay}")


# Calculate the total sevice delay of the NS-ESPs.
r_sur_delay=[]
for i in range(n):
    sur_delay = sur_users[i]*(2*sur_updown_delay[j]+sur_computing_delay[j])
    r_sur_delay.append(sur_delay)
print(f"total sevice delay of the NS-ESPs:{r_sur_delay}")

Min_H=sum(r_sch_delay)+sum(r_sur_delay)
print("minimum sevice delay: ",Min_H)
