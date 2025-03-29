import numpy as np
import pandas as pd
import random
import math
from openpyxl import load_workbook
from scipy.optimize import minimize


### CoEdge: Calculate the maximum economic benefit Max_U when 35-ESP and 55-ESP handle all user requests, 
### and also obtain the service pricing for each ESP.

#Take 35-ESP as an example.

# read dataset from task scheduling_35.xlsx
book=load_workbook(r'D:\my work\task scheduling_35.xlsx')
# Distinguish whether ESP requires scheduling or not.
schedule_sheet = book["RS"]
schedule_resource = []
sch_users = []
sch_local = []
surplus_sheet = book["NS"]
surplus_resource = []
sur_users = []
sur_local = []

sch_row_num = 2
while sch_row_num <= 11:
    schedule_resource.append(schedule_sheet.cell(row=sch_row_num, column=5).value)
    sch_users.append(schedule_sheet.cell(row=sch_row_num, column=4).value)
    sch_local.append(schedule_sheet.cell(row=sch_row_num, column=2).value)
    sch_row_num = sch_row_num + 1
sur_row_num = 2
while sur_row_num <= 46:
    surplus_resource.append(surplus_sheet.cell(row=sur_row_num, column=5).value)
    sur_users.append(surplus_sheet.cell(row=sur_row_num, column=4).value)
    sur_local.append(surplus_sheet.cell(row=sch_row_num, column=2).value)
    sur_row_num = sur_row_num + 1
m=len(schedule_resource)
n=len(surplus_resource)
a_m=m/(m+n)
a_n=n/(m+n)
# Link Transmission Rate
R= 0.04
# Unit service cost for user and unit maintenance cost
b_1=100
b_2=80
# Computing Latency
sch_computing_delay = [random.uniform(1,30) for _ in range(m)]
sur_computing_delay = [random.uniform(1,30) for _ in range(n)]
# Upload/Download Transmission Latency
sch_updown_delay = [random.uniform(1,30) for _ in range(m)]
sur_updown_delay = [random.uniform(1,30) for _ in range(n)]
#35-ESP scheduling matrix
df_r = pd.read_excel(r'D:\my work\scheduling matrix_35.xlsx', header=None)
r = df_r.values.tolist()


# Formulate the objective function and constraints, 
# which include labeling ESP u as completed if its service profit is nonnegative
# and updating the gu according to Powell if ESP u has not yet completed.
def objective_function(x):
   sch_E=0
   for i in range(m):
       sch_R=sch_local[i]*b_1
       sch_O=sch_local[i]*b_2+np.sum([x[j]*r[i][j] for j in range(n)])
       sch_E+=(sch_R-sch_O) 
   sur_E=0
   for i in range(n):
       sur_R=sur_users[i]*b_1+np.sum([x[i]*r[j][i] for j in range(m)])
       sur_O=sur_users[i]*b_2
       sur_E+=(sur_R-sur_O)
   E=-a_m*sch_E-a_n*sur_E
   return E

# Define the boundary conditions (non-negative constraints) of the variables.）
bounds = [(0, 10)] * n
 
# Number of iterations: 1000
# Convergence tolerance: Convergence is achieved when the relative change is less than 1%.
options = {'disp': True, 'maxiter': 1000, 'ftol': 1e-2}
# Optimization method: Powell
result = minimize(objective_function, [0]*n, method='Powell',
                 bounds=bounds, options=options)

if result.success:
   print(f"minimum value of the objective function: {result.fun}")
   print(f"optimal solution: {result.x}")
   df = pd.DataFrame(result.x).astype('float')
   df.to_excel(r'D:\my work\service price_35.xlsx', index=False, engine='openpyxl')
else:
   print("Optimization failed.")
   print("Reasons for failure: ", result.message)

# Calculate the weighted sevice profit of the RS-ESPs.
x_sch_E=[]
for i in range(m):
    sch_R=sch_local[i]*b_1
    sch_O=sch_local[i]*b_2+np.sum([result.x[j]*r[i][j] for j in range(n)])
    sch_E=(sch_R-sch_O)
    x_sch_E.append(sch_E)
print("weighted sevice profit of the RS-ESPs:",x_sch_E)
# Calculate the weighted sevice profit of the NS-ESPs.
x_sur_E=[]
for i in range(n):
    sur_R=sur_users[i]*b_1+np.sum([result.x[i]*r[j][i] for j in range(m)])
    sur_O=sur_users[i]*b_2
    sur_E=(sur_R-sur_O)
    x_sur_E.append(sur_E)
print("weighted sevice profit of the NS-ESPs",x_sur_E)
print("sevice profit of CoEdge: ",a_m*sum(x_sch_E)+a_n*sum(x_sur_E))
